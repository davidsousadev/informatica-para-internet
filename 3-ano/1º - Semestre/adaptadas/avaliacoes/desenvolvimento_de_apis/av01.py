from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import random, string

terms = [
    "API", "ENDPOINT", "REQUEST", "RESPONSE",
    "JSON", "XML", "GET", "POST",
    "PUT", "DELETE", "PATCH", "HEADER",
    "TOKEN", "AUTENTICACAO", "AUTORIZACAO",
    "REST", "SOAP", "CLIENTE",
    "SERVIDOR", "STATUS", "CACHE",
    "PAYLOAD", "PARAMETRO", "BODY",
    "JWT", "HTTPS", "URL", "COOKIE"
]

WIDTH = 14
HEIGHT = 24

grid = [["" for _ in range(WIDTH)] for _ in range(HEIGHT)]
dirs = [(1,0),(0,1),(1,1),(-1,1)]

random.seed()

for word in sorted(terms, key=len, reverse=True):
    placed = False

    for _ in range(1000):
        dx, dy = random.choice(dirs)

        x = random.randint(0, WIDTH - 1)
        y = random.randint(0, HEIGHT - 1)

        ex = x + dx * (len(word) - 1)
        ey = y + dy * (len(word) - 1)

        if not (0 <= ex < WIDTH and 0 <= ey < HEIGHT):
            continue

        ok = True

        for i, ch in enumerate(word):
            xx = x + dx * i
            yy = y + dy * i

            if grid[yy][xx] not in ("", ch):
                ok = False
                break

        if ok:
            for i, ch in enumerate(word):
                xx = x + dx * i
                yy = y + dy * i
                grid[yy][xx] = ch

            placed = True
            break

    if not placed:
        print(f"Não foi possível inserir: {word}")

for r in range(HEIGHT):
    for c in range(WIDTH):
        if grid[r][c] == "":
            grid[r][c] = random.choice(string.ascii_uppercase)

wb = Workbook()
ws = wb.active
ws.title = "Caça-Palavras"

ws["A1"] = "CAÇA-PALAVRAS - TERMOS BÁSICOS DE APIs"
ws["A1"].font = Font(bold=True)

for r in range(HEIGHT):
    for c in range(WIDTH):
        cell = ws.cell(row=r + 3, column=c + 1)
        cell.value = grid[r][c]
        cell.alignment = Alignment(horizontal="center")

start_col = WIDTH + 3

ws.cell(row=2, column=start_col, value="TERMOS")

for i, t in enumerate(terms, start=3):
    ws.cell(row=i, column=start_col, value=t)

path = "caca_palavras_termos_apis.xlsx"
wb.save(path)

print(path)